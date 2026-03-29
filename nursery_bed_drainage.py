#!/usr/bin/env python3
"""
苗床下水系统管材用量计算 Excel 生成脚本
- 苗床规格：4 ft × 40 ft（每单元 8 ft，共 5 单元）
- 可扩展：层数 × 组数
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── 颜色常量 ──────────────────────────────────────────────
CLR_HEADER      = "2E5A8E"   # 深蓝 - 表头
CLR_SECTION     = "4A90D9"   # 中蓝 - 分类标题
CLR_SUBHEADER   = "D9E8F5"   # 浅蓝 - 子表头
CLR_INPUT       = "FFF9C4"   # 淡黄 - 可输入单元格
CLR_RESULT      = "E8F5E9"   # 淡绿 - 计算结果
CLR_TOTAL       = "C8E6C9"   # 绿色 - 汇总行
CLR_WHITE       = "FFFFFF"
CLR_ALT         = "F5F9FF"   # 交替行背景

def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="2E5A8E")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, row, col_start, col_end, text, color=CLR_HEADER, font_size=12):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = header_font(size=font_size)
    cell.fill = fill(color)
    cell.alignment = center()
    cell.border = thick_border()

def apply_cell(ws, row, col, value, bg=CLR_WHITE, bold=False,
               align=None, num_format=None, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = body_font(bold=bold, color=color)
    c.fill = fill(bg)
    c.alignment = align or center()
    c.border = thin_border()
    if num_format:
        c.number_format = num_format
    return c

# ────────────────────────────────────────────────────────────────────────────
# 每个【苗床单元】(8ft) 所需管材（基础单元量）
# 4 × 40 苗床 = 5 单元，1 层，1 组
# ────────────────────────────────────────────────────────────────────────────
# 件名、规格、类型、每单元数量、单位、备注
PARTS = [
    # name,               spec,           part_type,          qty_per_unit, unit,  note
    ("T型三通接头",        '1"',           "T (Tee)",          4,            "个",  "三通分支接头；每5单元共用，按比例分配"),
    ("L型弯头",            '1"',           "Elbow 90°",        2,            "个",  "端部转角；每组固定2个"),
    ("直接头",             '1"',           "Coupling",         2,            "个",  "管道直线延伸连接；每组固定2个"),
    ("短管 11.5\"",        '1" × 11.5"',   "PVC Pipe",         2,            "条",  "端头过渡段"),
    ("短管 19\"",          '1" × 19"',     "PVC Pipe",         2,            "条",  "分支引出管"),
    ("长管 58\"",          '1" × 58"',     "PVC Pipe",         1,            "条",  "主横管段"),
    ("长管 59.5\"",        '1" × 59.5"',   "PVC Pipe",         1,            "条",  "主横管段"),
    ("长管 10'",           '1" × 10\'',    "PVC Pipe",         2,            "条",  "主干管（每单元1条）×2行"),
    ("镀锌管夹",           '1"',           "Galvanized Strap", 5,            "个",  "固定管道用"),
]

# 固定用量 vs 随单元线性扩展
# 分析：
#   - L型弯头(2)、直接头(2)、短管11.5"(2)、短管19"(2) → 每组固定，不随单元增加
#   - T型三通(4) → 每5单元/组4个，即0.8个/单元
#   - 长管 58"(1)、59.5"(1) → 每组固定各1条
#   - 长管 10'(2) → 每单元 2 条（线性）
#   - 镀锌管夹(5) → 每单元约1个（线性），每组5个
# 为方便扩展，我们定义：每组（5单元/一个完整苗床）的用量 = 基础量
# 当组数/层数变化时 = 基础量 × 层数 × 组数

def create_excel(output_path):
    wb = openpyxl.Workbook()

    # ── Sheet 1：参数配置 + 汇总 ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "管材用量汇总"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 10
    ws1.column_dimensions["G"].width = 28
    ws1.row_dimensions[1].height = 36
    ws1.row_dimensions[2].height = 18

    # 大标题
    apply_header_row(ws1, 1, 1, 7,
                     "🌱  苗床下水系统管材用量计算表", CLR_HEADER, 14)
    apply_header_row(ws1, 2, 1, 7,
                     "苗床规格：4 ft × 40 ft  |  每单元：8 ft  |  管径：1\"  |  每组单元数：5",
                     CLR_SECTION, 10)

    # ── 参数输入区 ──────────────────────────────────────────────────────────
    r = 4
    ws1.row_dimensions[r].height = 22
    apply_header_row(ws1, r, 1, 7, "▌ 参数配置（黄色单元格可修改）",
                     CLR_SECTION, 11)

    r += 1
    ws1.row_dimensions[r].height = 20
    params = [
        ("B", "参数项", CLR_SUBHEADER, True),
        ("C", "数值", CLR_SUBHEADER, True),
        ("D", "单位", CLR_SUBHEADER, True),
        ("E", "", CLR_SUBHEADER, True),
        ("F", "", CLR_SUBHEADER, True),
        ("G", "说明", CLR_SUBHEADER, True),
    ]
    for col_letter, text, bg, bold in params:
        col = openpyxl.utils.column_index_from_string(col_letter)
        c = ws1.cell(row=r, column=col, value=text)
        c.font = body_font(bold=bold, color="2E5A8E")
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()
    # A列序号列
    ws1.cell(row=r, column=1).fill = fill(CLR_SUBHEADER)
    ws1.cell(row=r, column=1).border = thin_border()

    param_data = [
        ("层数（Tiers）",     "=$C$6",  "层",  "苗床架的层数，默认1层"),
        ("组数（Sets）",      "=$C$7",  "组",  "苗床排列的组数，默认1组"),
        ("单元数/组（Units）","=$C$8",  "单元","每组单元数，固定5（每单元8ft）"),
    ]
    # 实际输入行
    r += 1  # row 6
    PARAM_ROW_START = r
    input_vals = [1, 1, 5]
    input_labels = ["层数（Tiers）", "组数（Sets）", "单元数/组（Units）"]
    input_units  = ["层", "组", "单元（8ft/单元）"]
    input_notes  = [
        "苗床架的层数，默认 1 层；增加层数按此倍增",
        "苗床排列的组数，默认 1 组（1个4×40苗床）",
        "每组单元数，固定 5（每单元 8 ft，共 40 ft）",
    ]
    for i, (label, val, unit, note) in enumerate(
            zip(input_labels, input_vals, input_units, input_notes)):
        cur_r = r + i
        ws1.row_dimensions[cur_r].height = 22
        # A - 序号
        apply_cell(ws1, cur_r, 1, i+1, CLR_WHITE, align=center())
        # B - 参数名
        apply_cell(ws1, cur_r, 2, label, CLR_WHITE, bold=True, align=left())
        # C - 输入值（黄色）
        c = ws1.cell(row=cur_r, column=3, value=val)
        c.font = body_font(bold=True, color="B8860B")
        c.fill = fill(CLR_INPUT)
        c.alignment = center()
        c.border = thick_border()
        c.number_format = "0"
        # D - 单位
        apply_cell(ws1, cur_r, 4, unit, CLR_WHITE, align=center())
        # E, F - 空
        apply_cell(ws1, cur_r, 5, "", CLR_WHITE)
        apply_cell(ws1, cur_r, 6, "", CLR_WHITE)
        # G - 说明
        apply_cell(ws1, cur_r, 7, note, CLR_WHITE, align=left())

    # 行号：6=层数, 7=组数, 8=单元数
    TIERS_CELL  = "C6"
    SETS_CELL   = "C7"
    UNITS_CELL  = "C8"

    r = PARAM_ROW_START + 3 + 1  # row 10

    # ── 管材清单表头 ────────────────────────────────────────────────────────
    ws1.row_dimensions[r].height = 22
    apply_header_row(ws1, r, 1, 7, "▌ 管材用量明细", CLR_SECTION, 11)
    r += 1  # row 11

    ws1.row_dimensions[r].height = 32
    col_headers = ["#", "零件名称", "规格", "类型",
                   "每组用量", "单位",
                   "总用量\n（层×组×每组）", "备注"]
    col_widths   = [5, 22, 16, 16, 10, 8, 16, 28]
    ws1.column_dimensions["H"] = ws1.column_dimensions.get("H") or \
        openpyxl.worksheet.dimensions.ColumnDimension(ws1, index="H")
    ws1.column_dimensions["H"].width = 28

    for ci, (hdr, w) in enumerate(zip(col_headers, col_widths), start=1):
        c = ws1.cell(row=r, column=ci, value=hdr)
        c.font = header_font(size=10)
        c.fill = fill(CLR_HEADER)
        c.alignment = center()
        c.border = thin_border()
        ws1.column_dimensions[get_column_letter(ci)].width = w

    r += 1  # row 12  ← 数据从这行开始
    DATA_ROW_START = r

    for i, (name, spec, ptype, qty, unit, note) in enumerate(PARTS):
        cur_r = r + i
        ws1.row_dimensions[cur_r].height = 20
        bg = CLR_WHITE if i % 2 == 0 else CLR_ALT
        apply_cell(ws1, cur_r, 1, i+1, bg, align=center())
        apply_cell(ws1, cur_r, 2, name, bg, bold=False, align=left())
        apply_cell(ws1, cur_r, 3, spec, bg, align=center())
        apply_cell(ws1, cur_r, 4, ptype, bg, align=center())
        apply_cell(ws1, cur_r, 5, qty, bg, align=center())
        apply_cell(ws1, cur_r, 6, unit, bg, align=center())
        # 总用量公式 = 每组用量 × 层数 × 组数
        total_formula = f"=E{cur_r}*{TIERS_CELL}*{SETS_CELL}"
        c_total = ws1.cell(row=cur_r, column=7, value=total_formula)
        c_total.font = body_font(bold=True, color="1A5E20")
        c_total.fill = fill(CLR_RESULT)
        c_total.alignment = center()
        c_total.border = thin_border()
        c_total.number_format = "0"
        apply_cell(ws1, cur_r, 8, note, bg, align=left())

    r += len(PARTS)  # row 21

    # ── 分类小计 ────────────────────────────────────────────────────────────
    ws1.row_dimensions[r].height = 22
    apply_header_row(ws1, r, 1, 8, "▌ 分类汇总", CLR_SECTION, 11)
    r += 1  # row 22

    # 管件小计
    # T型(row12) + L型(row13) + 直接头(row14) = 行 12~14
    FITTING_ROWS = list(range(DATA_ROW_START, DATA_ROW_START + 3))   # T, L, Coupling
    PIPE_ROWS    = list(range(DATA_ROW_START + 3, DATA_ROW_START + 8)) # 各管子
    STRAP_ROWS   = [DATA_ROW_START + 8]                                 # 管夹

    subtotals = [
        ("管件合计（T + L + Coupling）", FITTING_ROWS, "个"),
        ("管子合计（各规格 PVC 管）",    PIPE_ROWS,    "条"),
        ("管夹合计",                      STRAP_ROWS,   "个"),
    ]

    for j, (label, rows, unit) in enumerate(subtotals):
        cur_r = r + j
        ws1.row_dimensions[cur_r].height = 20
        apply_cell(ws1, cur_r, 1, "", CLR_TOTAL)
        apply_cell(ws1, cur_r, 2, label, CLR_TOTAL, bold=True, align=left())
        apply_cell(ws1, cur_r, 3, "", CLR_TOTAL)
        apply_cell(ws1, cur_r, 4, "", CLR_TOTAL)
        apply_cell(ws1, cur_r, 5, "", CLR_TOTAL)
        apply_cell(ws1, cur_r, 6, unit, CLR_TOTAL, align=center())
        sum_formula = "=" + "+".join([f"G{rr}" for rr in rows])
        c = ws1.cell(row=cur_r, column=7, value=sum_formula)
        c.font = body_font(bold=True, color="1A5E20")
        c.fill = fill(CLR_TOTAL)
        c.alignment = center()
        c.border = thin_border()
        c.number_format = "0"
        apply_cell(ws1, cur_r, 8, "", CLR_TOTAL)

    r += len(subtotals)

    # ── 注释 ────────────────────────────────────────────────────────────────
    r += 1
    ws1.row_dimensions[r].height = 18
    notes_text = ("💡 使用说明：修改第6行[层数]和第7行[组数]（黄色单元格），"
                  "所有用量将自动重新计算。单元数/组固定为5（4x40苗床）。")
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws1.cell(row=r, column=1, value=notes_text)
    c.font = Font(name="Arial", size=9, italic=True, color="555555")
    c.alignment = left()

    # ── Sheet 2：多场景对比表 ──────────────────────────────────────────────
    ws2 = wb.create_sheet("多场景对比")
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 22

    apply_header_row(ws2, 1, 1, 10,
                     "🌱  苗床下水系统 — 多场景用量对比", CLR_HEADER, 13)
    ws2.row_dimensions[1].height = 34

    # 场景定义：(层数, 组数)
    scenarios = [
        (1, 1),
        (1, 2),
        (1, 3),
        (2, 1),
        (2, 2),
        (2, 3),
        (3, 1),
        (3, 3),
    ]

    # 表头
    r2 = 3
    ws2.row_dimensions[r2].height = 30
    ws2.cell(row=r2, column=1).value = "#"
    ws2.cell(row=r2, column=1).font = header_font(size=10)
    ws2.cell(row=r2, column=1).fill = fill(CLR_HEADER)
    ws2.cell(row=r2, column=1).alignment = center()
    ws2.cell(row=r2, column=1).border = thin_border()
    ws2.cell(row=r2, column=2).value = "零件名称"
    ws2.cell(row=r2, column=2).font = header_font(size=10)
    ws2.cell(row=r2, column=2).fill = fill(CLR_HEADER)
    ws2.cell(row=r2, column=2).alignment = center()
    ws2.cell(row=r2, column=2).border = thin_border()

    for sci, (tiers, sets) in enumerate(scenarios):
        col = 3 + sci
        ws2.column_dimensions[get_column_letter(col)].width = 12
        label = f"{tiers}层×{sets}组\n({tiers*sets*5}单元)"
        c = ws2.cell(row=r2, column=col, value=label)
        c.font = header_font(size=9)
        c.fill = fill(CLR_HEADER)
        c.alignment = center()
        c.border = thin_border()

    r2 += 1
    for i, (name, spec, ptype, qty, unit, note) in enumerate(PARTS):
        cur_r2 = r2 + i
        ws2.row_dimensions[cur_r2].height = 20
        bg = CLR_WHITE if i % 2 == 0 else CLR_ALT
        ws2.cell(row=cur_r2, column=1, value=i+1).fill = fill(bg)
        ws2.cell(row=cur_r2, column=1).alignment = center()
        ws2.cell(row=cur_r2, column=1).border = thin_border()
        ws2.cell(row=cur_r2, column=2, value=f"{name} ({spec})").fill = fill(bg)
        ws2.cell(row=cur_r2, column=2).alignment = left()
        ws2.cell(row=cur_r2, column=2).border = thin_border()
        ws2.cell(row=cur_r2, column=2).font = body_font()

        for sci, (tiers, sets) in enumerate(scenarios):
            col = 3 + sci
            total = qty * tiers * sets
            c = ws2.cell(row=cur_r2, column=col, value=total)
            c.font = body_font(bold=True if total > qty else False,
                               color="1A5E20" if total > qty else "000000")
            c.fill = fill(CLR_RESULT if total > qty else bg)
            c.alignment = center()
            c.border = thin_border()
            c.number_format = "0"

    # 保存
    wb.save(output_path)
    print(f"✅ Excel 已生成：{output_path}")

if __name__ == "__main__":
    out = "/Users/scotpan/WorkBuddy/20260329080542/苗床下水系统管材用量表.xlsx"
    create_excel(out)
